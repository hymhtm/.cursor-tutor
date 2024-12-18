import os
import time
import openpyxl
from openpyxl.formatting.rule import Rule, DataBarRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd
from config import settings
from resolve_path import resolve_path

def pivot_styler(xls_file_path):
    #ファイルパスと拡張子を確認
    path = resolve_path(xls_file_path)
    if path.split('.')[-1] != 'xlsx':
        raise ValueError("このファイルは操作できません")
    else:
        abs_path = path


    #エクセルファイルを開く
    wb = openpyxl.load_workbook(abs_path)
    ws = wb.active

    #データ範囲を取得
    last_row_num = ws.max_row
    last_col_num = ws.max_column
    last_col_letter = get_column_letter(last_col_num)

    #スタイルと列幅設定
    for col_ in range(2, last_col_num+1):
        col = ws.column_dimensions[get_column_letter(col_)]
        col.width = 17
        for cell in ws[col_][1:last_row_num+1]:
            cell.number_format = '0.0%'
            cell.font = Font(name='Yu Gothic UI', size=11)

    #行高さ設定
    for row in range(1, last_row_num+1):
        ws.row_dimensions[row].height = 18.75


    #データバーの条件付き書式を設定
    bar_rule = DataBarRule(
        start_type='percent',
        start_value=0,
        end_type='percent',
        end_value=100,
        color='00efab',
        showValue=True
    )

    #データバーの条件付き書式を適用
    ws.conditional_formatting.add('B2:{}{}'.format(last_col_letter, last_row_num), bar_rule)
    


    #A列の日付の書式設定
    for cell in ws['A'][1:last_row_num+1]:
        cell.number_format = 'm月d日'

    wb.save(abs_path)

if __name__ == '__main__':
    print("start")
    start_time = time.time()
    pivot_styler(r"C:\Users\nakamura114\Downloads\Report\Report\report_output_Nov\monthly_pivot_table.xlsx")
    end_time = time.time()
    print(f"end. time:{end_time - start_time}sec")